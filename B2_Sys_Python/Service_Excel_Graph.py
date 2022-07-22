# region Import
# region Python
import openpyxl
from openpyxl.chart import BarChart,Reference
# endregion

# region Other
import sys
import pandas as pd
import datetime
try:
    sys.path.append('S:/Sahu_Group/B3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/B2_Sys_Python')  
    # print(sys.path)
    import Service_Excel
    from Service_Excel import *
except ModuleNotFoundError as e:
    print(e)
    print('Absolute import failed')
# endregion
# endregion

# region Class-Service_Excel_Chart
class Service_Excel_Chart_Class:
    @staticmethod
    def Create_Chart_Bar(GraphModel,Excel_Worksheet):
        # Create object of BarChart class
        chart = BarChart()

        # write o to 9 in 1st column of the active sheet
        for i in range(0,len(GraphModel["Values"])):
        	Excel_Worksheet.append([ GraphModel["Values"][i]  ])

        Values = Reference(Excel_Worksheet, min_col = 1, min_row = 1,max_col = 1, max_row = 10)

        # adding data to the Bar chart object
        chart.add_data(Values)

        # set the title of the chart
        chart.title = GraphModel["Title"]

        # set the title of the x-axis
        chart.x_axis.title = GraphModel["TitleX"]

        # set the title of the y-axis
        chart.y_axis.title = GraphModel["TitleY"]

        Excel_Worksheet.add_chart(chart,GraphModel["StartPoint"])

    @staticmethod
    def Create_Chart(GraphModel,File_Path,Sheet_Index):
        Service_Excel_Chart_Class_Var = Service_Excel_Chart_Class()
        Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
        Excel_Worksheet_Name = "Sheet"+str(Sheet_Index+1)
        Excel_Worksheet = None

        try:
            if Read_Check_SheetPresentOrNot_Excel_Workbook(Excel_Workbook,Excel_Worksheet_Name) == True:
                Excel_Workbook = OpenPyXL_Delete_Worksheet(Excel_Workbook,OpenPyXL_Open_Excel_Workbook_Worksheet_ByName(Excel_Workbook,Excel_Worksheet_Name))
                Excel_Workbook = OpenPyXL_Update_AddSheet_Excel_Workbook(Excel_Workbook,Excel_Worksheet_Name)
            else:
                Excel_Workbook = OpenPyXL_Update_AddSheet_Excel_Workbook(Excel_Workbook,Excel_Worksheet_Name)

            Excel_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByName(Excel_Workbook,Excel_Worksheet_Name)
            Excel_Worksheet = OpenPyXL_Update_Clear_AllRows_Worksheet(Excel_Worksheet)
        except:
            pass
        # ---
        if GraphModel["GraphName"]=="Bar":
            Excel_Worksheet = Service_Excel_Chart_Class_Var.Create_Chart_Bar(GraphModel,Excel_Worksheet)
        # ---
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)  
# endregion

# region Declare-Var
GraphModel ={
    "GraphName":"Bar",
    "Title":"CHART-BAR",
    "TitleX":"X_AXIS",
    "TitleY":"Y_AXIS",
    "StartPoint": "A1",
    "Values": [1,2,3]
}
# endregion

# region Test
File_Path = "S:/Sahu_Group/B3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/C1_Out/ExcelGraph/Chart_Bar_100.xlsx"
Sheet_Index = 0
Service_Excel_Chart_Class_Var = Service_Excel_Chart_Class()
Excel_Workbook = Service_Excel_Chart_Class_Var.Create_Chart(GraphModel,File_Path,Sheet_Index)
# endregion
