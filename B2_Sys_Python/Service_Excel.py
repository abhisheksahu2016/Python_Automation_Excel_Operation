# ---
# region Import
from decimal import Context
import os
# ---
import numpy as np # [Info]:Numpy for .xls./Panda for .csv
from numpy import * 
# ---
import xlrd 
import xlwt # [Info]:Doesn't Support Xlsx
import xlsxwriter # [Info]:Doesn't Support Update
import openpyxl # [Info]:Index Start With 1
# endregion 
# ---
Excel_Expt_Fldr_In_Path = "S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In"
Excel_Expt_File_In_Name_100 = "Default.xlsx"

Excel_Expt_Fldr_Out_Path = "S:/Sahu_Group/C3_Data/C2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/C1_Out"
Excel_Expt_File_Out_Name_100 = "Default.xlsx"
# ---
# print(Get_File_Path("In","100"))
def Get_File_Path(Fldr_Name,Option):
    if Option=="100" and Fldr_Name=="In" :
        return Excel_Expt_Fldr_In_Path + "/" + Excel_Expt_File_In_Name_100
    if Option=="100" and Fldr_Name=="Out" :
        return Excel_Expt_Fldr_Out_Path + "/Default/" + Excel_Expt_File_Out_Name_100
# ---
# {Create}-Excel_Workbook
def XLWT_Create_Excel_Workbook_Blank():
    Excel_Workbook = xlwt.Workbook()
    return Excel_Workbook
# ---
def XLWT_Create_Excel_Workbook_FromPath_Default():
    Excel_Workbook = xlwt.Workbook(Get_File_Path("In","100"))
    return Excel_Workbook
# ---
def XLWT_Create_Excel_Workbook_FromPath_Mannual(File_Path):
    Excel_Workbook = xlwt.Workbook(File_Path)
    return Excel_Workbook
# ---
def XLSXWriter_Create_Excel_Workbook_Blank():
    Excel_Workbook = xlsxwriter.Workbook()
    return Excel_Workbook
# --
def XLSXWriter_Create_Excel_Workbook_FromPath_Default():
    Excel_Workbook = xlsxwriter.Workbook(Get_File_Path("In","100"))
    return Excel_Workbook
# ---
def XLSXWriter_Create_Excel_Workbook_FromPath_Mannual(File_Path):
    Chk_File = os.path.exists(File_Path) 
    try:        
        if Chk_File is True :
            Excel_Workbook = xlsxwriter.Workbook(File_Path)
            print(File_Path)
            print('File Succesfully OverWrited')
        else :
            Excel_Workbook = xlsxwriter.Workbook(File_Path)
            print(File_Path)
            print('File Succesfully NewWrited')
    except:
        print("Error Came")
    return Excel_Workbook
# ---
def XLWT_Create_Excel_Workbook_Blank():
    Excel_Workbook = openpyxl.Workbook()
    return Excel_Workbook
# ---
def OpenPyXL_Create_Excel_Workbook_FromPath_Mannual(File_Path):
    Chk_File = os.path.exists(File_Path) 
    try:        
        if Chk_File is True :
            Excel_Workbook = openpyxl.Workbook(File_Path)
            print(File_Path)
            print('File Succesfully OverWrited')
        else :
            Excel_Workbook = openpyxl.Workbook(File_Path)
            print(File_Path)
            print('File Succesfully NewWrited')
    except:
        print("Error Came")
    return Excel_Workbook
# ---
def OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path):
    Chk_File = os.path.exists(File_Path) 
    Function_Message = ""
    try:        
        if Chk_File is True :
            Excel_Workbook = Excel_Workbook.save(File_Path)
            print(File_Path)
            Function_Message = "Success:OpenPyXL_Save_Excel_Workbook_FromPath_Mannual"
            print('File Succesfully OverWrited')
            return Excel_Workbook,Function_Message
        else :
            Excel_Workbook = Excel_Workbook.save(File_Path)
            print(File_Path)
            Function_Message = "Success:OpenPyXL_Save_Excel_Workbook_FromPath_Mannual"
            print('File Succesfully NewWrited')
            return Excel_Workbook,Function_Message
    except:
        print("Error Came")
        Function_Message = "Failure:OpenPyXL_Save_Excel_Workbook_FromPath_Mannual"
        return Excel_Workbook,Function_Message

# ---
# {Create}-Excel_Workbook-Format
def XLSXWriter_Create_Excel_Workbook_Format(Excel_Workbook,FormatType):
    Excel_Workbook_Format = Excel_Workbook.add_format()
    # ------------
    FormatType_Split = FormatType.split("-")
    # ------------
    for i in range(0,len(FormatType_Split)):
        FormatType_X = FormatType_Split[i].split(":")[0]
        FormatType_Y = FormatType_Split[i].split(":")[1]
        # print(FormatType_X)
        # print(FormatType_Y)
        # Cell-Content-Format-Font/Font-Color/Font-Pattern
        # Cell-Content-Format-Number To Currency
        if FormatType_X.find("NumFormat")!=-1:
            if FormatType_Y.find("Currency")!=-1:
                Excel_Workbook_Format = Excel_Workbook.add_format({'num_format': '$#,##0.00'})
        if FormatType_X.find("FontBold")!=-1:
            Excel_Workbook_Format.set_bold(False)
            if FormatType_Y.find("True")!=-1:
                Excel_Workbook_Format.set_bold(bool(FormatType_Y))
        if FormatType_X.find("FontColor")!=-1:
            Excel_Workbook_Format.set_font_color(FormatType_Y)
        if FormatType_X.find("BgColor")!=-1:
            Excel_Workbook_Format.set_bg_color(FormatType_Y)
        if FormatType_X.find("Border")!=-1:
            Excel_Workbook_Format.set_border(int(FormatType_Y))

    # ------------
    return Excel_Workbook_Format
# ---
def XLSXWriter_Create_Excel_Workbook_Worksheet_Format(Excel_Workbook_Worksheet,Excel_Workbook_Worksheet_Rows,Excel_Workbook_Worksheet_Cols,FormatType):
    if isinstance(Excel_Workbook_Worksheet_Rows, list) is True :
        Excel_Workbook_Worksheet.set_row(Excel_Workbook_Worksheet_Rows[0],Excel_Workbook_Worksheet_Rows[1], FormatType)
    if isinstance(Excel_Workbook_Worksheet_Cols, str) is True and Excel_Workbook_Worksheet_Cols.find("NA")==-1:
        Excel_Workbook_Worksheet.set_column('A:D', 20, FormatType)
    return Excel_Workbook_Worksheet
# {Create}-Excel_Workbook_Worksheet
# ---
def XLWT_Create_Excel_Workbook_Worksheet_Blank(Excel_Workbook):
    Excel_Workbook_Worksheet = Excel_Workbook.add_sheet()
    return Excel_Workbook,Excel_Workbook_Worksheet
def XLWT_Create_Excel_Workbook_Worksheet_Default(Excel_Workbook):
    Excel_Workbook_Worksheet = Excel_Workbook.add_sheet("Sheet1",cell_overwrite_ok=True)
    return Excel_Workbook,Excel_Workbook_Worksheet
# ---
def XLWT_Create_Excel_Workbook_Worksheet_Manual(Excel_Workbook,Excel_Workbook_WorksheetName):
    Excel_Workbook_Worksheet = Excel_Workbook.add_sheet(Excel_Workbook_WorksheetName,cell_overwrite_ok=True)
    return Excel_Workbook,Excel_Workbook_Worksheet
# ---
def XLSXWriter_Create_Excel_Workbook_Worksheet_Blank(Excel_Workbook):
    Excel_Workbook_Worksheet = Excel_Workbook.add_worksheet()
    return Excel_Workbook_Worksheet
# ---
def XLSXWriter_Create_Excel_Workbook_Worksheet_Default(Excel_Workbook):
    Excel_Workbook_Worksheet = Excel_Workbook.add_worksheet("Sheet1")
    return Excel_Workbook_Worksheet
# ---
def XLSXWriter_Create_Excel_Workbook_Worksheet_Mannual(Excel_Workbook,Excel_Workbook_WorksheetName):
    Excel_Workbook_Worksheet = Excel_Workbook.add_worksheet(Excel_Workbook_WorksheetName)
    return Excel_Workbook_Worksheet
# ---
def OpenPyXL_Create_Excel_Workbook_Worksheet_Mannual(Excel_Workbook,Excel_Workbook_WorksheetIndex,Excel_Workbook_WorksheetName):
    Excel_Workbook_Worksheet = Excel_Workbook.create_sheet(index = Excel_Workbook_WorksheetIndex,title=Excel_Workbook_WorksheetName)
    return Excel_Workbook_Worksheet

def OpenPyXL_Update_Clear_AllRows_Worksheet(Excel_Workbook_Worksheet):
    Excel_Workbook_Worksheet.delete_rows(1, Excel_Workbook_Worksheet.max_row+1) 
    return Excel_Workbook_Worksheet

def OpenPyXL_Delete_Worksheet(Excel_Workbook,Excel_Workbook_Worksheet):
    Excel_Workbook.remove_sheet(Excel_Workbook_Worksheet)
    return Excel_Workbook

def OpenPyXL_Update_AddSheet_Excel_Workbook(Excel_Workbook,Excel_Workbook_Worksheet_Name):
    Excel_Workbook.create_sheet(Excel_Workbook_Worksheet_Name)
    return Excel_Workbook

def Read_Check_SheetPresentOrNot_Excel_Workbook(Excel_Workbook_Workbbok,Excel_Workbook_Worksheet_Name):
    if Excel_Workbook_Worksheet_Name in Excel_Workbook_Workbbok.sheetnames:
        return True
    else:
        return False

# {Create}-Excel
def XLWT_Create_Excel_FromWorkbook(Excel_Workbook,File_Path):
    Chk_File = os.path.exists(File_Path) 
    if Chk_File is True :
        Excel_Workbook.save(File_Path) 
        print(File_Path)
        print('File Succesfully OverWrited')
    else :
        with open(os.path.join(File_Path), 'w') as fp: 
            Excel_Workbook.save(File_Path) 
            print(File_Path)
            print('File Succesfully NewWrited')
    return Excel_Workbook
# ---
def XLSXWriter_Create_Excel_FromWorkbook(Excel_Workbook,File_Path):
    pass
# ---
# {Update-Write}-Excel_Workbook_Worksheet_Cell-{Content-Format}
def XLSXWriter_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index_Start,Content,CellFormat,Direction):
    # Check-Matrix
    if all(isinstance(ele, list) for ele in Content)==True:
        # Check-List
        if Direction=="H":
            for y in range(0,len(Content)):
                for x in range(0,len(Content[0])):
                    Excel_Workbook_Worksheet.write(Index_Start[0]+y,Index_Start[1]+x,str(Content[y][x]),CellFormat) 
        if Direction=="V":
            for x in range(0,len(Content[0])):
                for y in range(0,len(Content)):
                    Excel_Workbook_Worksheet.write(Index_Start[0]+y,Index_Start[1]+x,str(Content[y][x]),CellFormat) 
    else:
        # Check-List
        if len(Content):
            if Direction=="H":
                for i in range(0,len(Content)):
                    Excel_Workbook_Worksheet.write(Index_Start[0],Index_Start[1]+i,str(Content[i]),CellFormat) 

            if Direction=="V":
                for i in range(0,len(Content)):
                    Excel_Workbook_Worksheet.write(Index_Start[0]+i,Index_Start[1],str(Content[i]),CellFormat) 
        # Check-Point
        else:
            Excel_Workbook_Worksheet = Excel_Workbook_Worksheet.write(Index_Start[0],[Index_Start[1]],Content,CellFormat)
    return Excel_Workbook_Worksheet   
def XLWT_Write_Excel_Workbook_Worksheet_FromMatrix(Excel_Workbook_Worksheet,Index_Start,Excel_Workbook_Worksheet_Content,Direction):
    if Direction=="RowWise":
        Excel_Workbook_Worksheet_Content_YLen = len(Excel_Workbook_Worksheet_Content)
        Excel_Workbook_Worksheet_Content_XLen = len(Excel_Workbook_Worksheet_Content[0])

        Y_End = Index_Start[0] + Excel_Workbook_Worksheet_Content_YLen + 1 
        X_End = Index_Start[1] + Excel_Workbook_Worksheet_Content_XLen + 1 

        Content_Structure_Matrix=[]
        temp_str = " "
        for x in range(Index_Start[0],X_End):
            for y in range(Index_Start[1],Y_End):
                temp_str = Excel_Workbook_Worksheet_Content[y][x]
                Excel_Workbook_Worksheet.write(y,x, temp_str) 
        return Excel_Workbook_Worksheet
    if Direction=="ColWise":
        Excel_Workbook_Worksheet_Content_YLen = len(Excel_Workbook_Worksheet_Content)
        Excel_Workbook_Worksheet_Content_XLen = len(Excel_Workbook_Worksheet_Content[0])

        Y_End = Index_Start[0] + Excel_Workbook_Worksheet_Content_YLen + 1 
        X_End = Index_Start[1] + Excel_Workbook_Worksheet_Content_XLen + 1 

        Content_Structure_Matrix=[]
        temp_str = " "
        for y in range(Index_Start[1],Y_End):
            for x in range(Index_Start[0],X_End):
                temp_str = Excel_Workbook_Worksheet_Content[y][x]
                Excel_Workbook_Worksheet.write(y,x, temp_str) 
        return Excel_Workbook_Worksheet
# ---
def XLWT_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index,Excel_Workbook_Worksheet_Content,Direction):
    if Direction=="H":
        for i in range(0,len(Excel_Workbook_Worksheet_Content)):
            Excel_Workbook_Worksheet.write(Index[0],Index[1]+i,str(Excel_Workbook_Worksheet_Content[i])) 

    if Direction=="V":
        for i in range(0,len(Excel_Workbook_Worksheet_Content)):
            Excel_Workbook_Worksheet.write(Index[0]+i,Index[1],str(Excel_Workbook_Worksheet_Content[i])) 

    return Excel_Workbook_Worksheet
# ---
# OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index_Start,[1],CellFormat,Direction):
# OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index_Start,[1.2.3],CellFormat,Direction):
# OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index_Start,[[[1,2],[3,4]]],CellFormat,Direction):
    
def OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Index_Start,Content,CellFormat,Direction):
    # Check-Matrix
    if len(Content)>0 and isinstance(Content[0], list) is True:
    # if all(isinstance(ele, list) for ele in Content)==True:
    #     # Check-List
        if Direction=="H":
            for y in range(0,len(Content)):
                for x in range(0,len(Content[0])):
                    # Format
                    # Excel_Workbook_Worksheet.cell.fill = PatternFill(start_color=yellow, end_color=yellow,fill_type = "solid")
                    # Update
                    Excel_Workbook_Worksheet.cell(row=Index_Start[0]+y,column=Index_Start[1]+x).value = str(Content[y][x])

        if Direction=="V":
            for x in range(0,len(Content[0])):
                for y in range(0,len(Content)):
                    Excel_Workbook_Worksheet.cell(row=Index_Start[0]+y,column=Index_Start[1]+x).value = str(Content[y][x])
    else:
        # Check-List
        if len(Content):
            if Direction=="H":
                for i in range(0,len(Content)):
                    Excel_Workbook_Worksheet.cell(row=Index_Start[0],column=Index_Start[1]+i).value = str(Content[i])

            if Direction=="V":
                for i in range(0,len(Content)):
                    Excel_Workbook_Worksheet.cell(row=Index_Start[0]+i,column=Index_Start[1]).value = str(Content[i])
        # Check-Point
        else:
            Excel_Workbook_Worksheet = Excel_Workbook_Worksheet.cell(row=Index_Start[0],column=Index_Start[1]).value = Content
    return Excel_Workbook_Worksheet   
# ---
def Get_SerialToDeserial(Serial_Str):
    v = Serial_Str
    # ---
    # v = "Student" 
    # v = "Student-xxx"
    # v = "Student-{xxx}"
    # v = "{xxx}-Student"
    # v = "{xxx}-{Student}"
    # v = "{xxx}-{Student}"
    # v = "Student-{xxx}"
    # v = "Student-{xxx-ccc}"
    # v = "Student-{xxx-{ccc}}"
    # v = "Student-{{xxx}-ccc}"
    # v = "Student-{{xxx}-{ccc}}"
    # v = "Student-{{xxx}-ccc-{j}}"
    # v = "Student-{Time-Fuller}-{Service-Taker}-{School-Private-NA}-{Infromation-Education}-Educatie-{Mrimary-Aaaa}-{0-1-2}-NA"
    V_Word_List = []
    Index = 0
    while(Index < len(v)):
        # print(str(len(v))+"-"+str(Index))
        # print(str(Index)+"-"+v[Index])
        if v[Index]!="{":
            for Index_Temp in range(Index,len(v)):
                if v[Index_Temp]=="-":
                    # print(v[Index:Index_Temp:1])
                    V_Word_List.append(v[Index:Index_Temp:1])
                    Index = Index_Temp + 1       
                    break
                if Index_Temp==len(v)-1:
                    # print(v[Index_Temp:len(v):1])
                    V_Word_List.append(v[Index:Index_Temp+1:1])
                    Index = Index_Temp + 1     
                    break     
        elif v[Index]=="{":
            Left_Curl_Status_Count = 1
            # print(v[Index+1])
            for Index_Temp in range(Index+1,len(v)):
                if v[Index_Temp]=="{":
                    Left_Curl_Status_Count = Left_Curl_Status_Count + 1
                if v[Index_Temp]=="}":
                    Left_Curl_Status_Count = Left_Curl_Status_Count - 1
                if Left_Curl_Status_Count==0:
                    V_Word_List.append(v[Index:Index_Temp+1:1])
                    Index = Index_Temp + 2     
                    break     
    return V_Word_List
# ---
def OpenPyXL_Update_Excel_SerialDeSerialCellMatrix(File_Path,Excel_Workbook_Worksheet_Index):
    Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
    Excel_Workbook_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index)

    Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[0]
    Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[1]

    if Total_Col>1:#Work_Type == "Serialize":
        Content_List = []
        for y in range(1,Total_Row+1):
            Temp_Row_Str = ""
            for x in range(1,Total_Col+1):
                Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y,x],[y,x])[0]
                if Cell_Val=='None':
                    for y1 in range(y,0,-1):
                        Cell_Val_Upper = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y1,x],[y1,x])[0]
                        if Cell_Val_Upper != 'None':
                            Cell_Val = Cell_Val_Upper
                            break
                if x==1:
                    Temp_Row_Str = Temp_Row_Str + Cell_Val  
                else:
                    Temp_Row_Str = Temp_Row_Str + "-" + Cell_Val
            Content_List.append(Temp_Row_Str) 

        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[1,Total_Col+1],Content_List,None,"V")
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)
    # if Work_Type == "DeSerialize":
    #     Content_List = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[1,1],[Total_Row,1])
    #     # ---
    #     Content_Matrix_RowWise = []
    #     for y in range(0,Total_Row):
    #         Content_List_Split = Content_List[y].split("-")
    #         Content_Matrix_RowWise.append(Content_List_Split)
    #     # ---
    #     Content_Matrix_ColWise = []
    #     for x in range(0,len(Content_Matrix_RowWise[0])):
    #         Temp_List = []
    #         for y in range(0,len(Content_Matrix_RowWise)):
    #             Temp_List.append(Content_Matrix_RowWise[y][x])
    #         Content_Matrix_ColWise.append(Temp_List)
    #     # ---
    #     Content_Matrix_ColWise_Modified = []
    #     for x in range(0,len(Content_Matrix_ColWise)):
    #         Content_Matrix_ColWise_OneCol_Data_List_Unique = list(set(Content_Matrix_ColWise[x]))
    #         Content_Matrix_ColWise_OneCol_Data_List_Track = [False for x in range(0,len(Content_Matrix_ColWise_OneCol_Data_List_Unique))]        
    #         for y in range(0,Total_Row):
    #             Cell_Value = Content_Matrix_ColWise[y][x]
    #             if Cell_Value in Content_Matrix_ColWise_OneCol_Data_List_Unique and Content_Matrix_ColWise_OneCol_Data_List_Track[Content_Matrix_ColWise_OneCol_Data_List_Unique.index(Cell_Value)]!=True:
    #                 Content_Matrix_ColWise_OneCol_Data_List_Track[Content_Matrix_ColWise_OneCol_Data_List_Unique.index(Cell_Value)]=True
    #             else:
    #                 Content_Matrix_ColWise[y][x] = ""
    #     # ---
    #     OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,Content_Matrix_ColWise_Modified,"V")

    if Total_Col==1 : #Work_Type == "DeSerialize":
        Content_List = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[1,1],[Total_Row,1])
        for Row in range (0,len(Content_List)):
            V_Word_List = Get_SerialToDeserial(Content_List[Row])
            OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[Row+1,Total_Col+1],V_Word_List,None,"H")
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)
        # ---

def OpenPyXL_TableDescionSerialDeSerial(File_Path,Excel_Workbook_Worksheet_Index,Work_Type):
    Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
    Excel_Workbook_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index)

    Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[0]
    Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[1]
    # --
    Content_List = []
    for y in range(1,Total_Row+1):
        Temp_Row = []
        for x in range(1,Total_Col+1):
            Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y,x],[y,x])[0]
            Temp_Row.append(Cell_Val)
        Content_List.append(Temp_Row) 
    # --
    Modified_Content_List = []
    if Total_Col==1:# Work_Type == "Serialize":
        Table_Name = Content_List[0][0]
        Modified_Content_List.append(Table_Name)
        for EachRow in range(0,len(Content_List)):
            for EachCol in range(1,len(Content_List[EachRow])):
                Modified_Content_List.append(Content_List[EachRow][0]+"-"+Content_List[0][EachCol]+"-"+Content_List[EachRow][EachCol])
        Excel_Workbook_Worksheet2 = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index+1)
        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet2,[1,1],Modified_Content_List,None,"V")
    if Total_Col>1: # Work_Type == "DeSerialize":
        Table_Name = Content_List[0][0]
        # ---
        X_List = []
        Y_List = []

        for EachRow in range(0,len(Content_List)):
            X_List.append(Content_List[EachRow].split("-")[1])
            Y_List.append(Content_List[EachRow].split("-")[0])
        # ---
        X_List_New = []
        Y_List_New = []
        Status_List_New = []

        X_List_New =[X_List_New.append(x) for x in X_List if x not in X_List_New]
        Y_List_New =[Y_List_New.append(x) for x in Y_List if x not in Y_List_New]

        for EachRow in range(0,len(Y_List_New)+1):
            Temp_Row = []
            if EachRow==0:
                Temp_Row.append(Table_Name)
                for EachCol in range(0,len(X_List_New)):
                    Temp_Row.append(Content_List[EachRow])
                Modified_Content_List.append(Temp_Row)
            else:
                Temp_Row.append(Y_List_New[EachRow])
                for EachCol in range(0,len(X_List_New)+1):
                    # Find-Status
                    status = ""
                    for EachRow2 in range(0,len(Content_List)):
                        if Content_List[EachRow2].find(Y_List_New[EachRow])!=-1 and Content_List[EachRow2].find(X_List[EachCol])!=-1:
                            status = Content_List[EachRow].split("-")[2]
                    Temp_Row.append(status)
                Modified_Content_List.append(Temp_Row)
        Excel_Workbook_Worksheet2 = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index+1)
        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet2,[1,1],Modified_Content_List,None,"H")
    # --
    OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)
# ---
def OpenPyXL_Update_Excel_FillDeFillCellMatrix(File_Path,Excel_Workbook_Worksheet_Index,Work_Type):
    Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
    Excel_Workbook_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index)

    Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[0]
    Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[1]

    if Work_Type == "Fill":
        Content_Matrix = []
        for y in range(1,Total_Row+1):
            Temp_Row = []
            for x in range(1,Total_Col+1):
                Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y,x],[y,x])[0]
                Temp_Row.append(Cell_Val)
            Content_Matrix.append(Temp_Row)
        
        for row in range(1,len(Content_Matrix)):
            for col in range(0,len(Content_Matrix[row])):           
                if Content_Matrix[row][col]=='None' or Content_Matrix[row][col]=="":
                    Content_Matrix[row][col] = Content_Matrix[row-1][col]
    
        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[1,1],Content_Matrix,None,"H")
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)
    if Work_Type == "DeFill":
        Content_Matrix = []
        for y in range(1,Total_Row+1):
            Temp_Row = []
            for x in range(1,Total_Col+1):
                Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y,x],[y,x])[0]
                Temp_Row.append(Cell_Val)
            Content_Matrix.append(Temp_Row)
        
        for row in range(len(Content_Matrix)-1,0,-1):
            for col in range(len(Content_Matrix[row])-1,-1,-1):           
                if Content_Matrix[row][col]==Content_Matrix[row-1][col]:
                    Content_Matrix[row][col] = ""
    
        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[1,1],Content_Matrix,None,"H")
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)
    
        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[1,1],Content_Matrix,None,"H")
        OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)

# ---
def OpenPyXL_Update_Excel_KunchiKunchiCell(File_Path,Excel_Workbook_Worksheet_Index,Work_Style):
    Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
    Excel_Workbook_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Excel_Workbook_Worksheet_Index)

    Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[0]
    Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet)[1]

    Content_List = []
    for y in range(1,Total_Row+1):
        for x in range(1,Total_Col+1):
            Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,[y,x],[y,x])[0]
            if Cell_Val.find("-")==-1:
               Content_List.append(Cell_Val) 
            if Cell_Val.find("-")!=-1:
               Content_List.append(Work_Style[0]+Cell_Val+Work_Style[1]) 
    OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[1,Total_Col+1],Content_List,None,"V")
    OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)

def OpenPyXL_Update_Excel_IDReplaceCellMatrix(File_Path,Excel_Workbook_Worksheet_Index_Dict):
    #
    Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)

    for Total_Change in range(0,len(Excel_Workbook_Worksheet_Index_Dict)):        
        #
        Source_Index = Excel_Workbook_Worksheet_Index_Dict[Total_Change][0]
        
        Excel_Workbook_Worksheet_Source = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Source_Index)
        Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet_Source)[0]
        Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet_Source)[1]

        Source_Content_Dict = []

        for y in range(2,Total_Row+1):
            Cell_Val_Category = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet_Source,[y,1],[y,1])[0]
            Cell_Val_CategoryId = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet_Source,[y,2],[y,2])[0]
    
            Temp_Source_Content_Dict = []
            Temp_Source_Content_Dict.append(Cell_Val_Category)
            Temp_Source_Content_Dict.append(Cell_Val_CategoryId)
            
            Source_Content_Dict.append(Temp_Source_Content_Dict) 
        #         
        Destination_Index = Excel_Workbook_Worksheet_Index_Dict[Total_Change][1]

        Excel_Workbook_Worksheet_Destination = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Destination_Index)
        Total_Row = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet_Destination)[0]
        Total_Col = OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet_Destination)[1]

        Destination_Content_Matrix = []

        for y in range(1,Total_Row+1):
            Temp_Row = []
            for x in range(1,Total_Col+1):
                Cell_Val = OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet_Destination,[y,x],[y,x])[0]
                Temp_Row.append(Cell_Val) 
            Destination_Content_Matrix.append(Temp_Row)

        for Row in range(0,len(Destination_Content_Matrix)):
            for Col in range(0,len(Destination_Content_Matrix[Row])):
                for Id in range(0,len(Source_Content_Dict)):
                    if Source_Content_Dict[Id][0] == Destination_Content_Matrix[Row][Col]:
                        Destination_Content_Matrix[Row][Col] = Source_Content_Dict[Id][1]

        OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet_Destination,[1,1],Destination_Content_Matrix,None,"H")
        # OpenPyXL_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet_Destination,[1,1],[[4, 5], [5, 8], [9, 10]],None,"H")
    
    OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)

# {Read-Open}-Excel_Workbook
def XLRD_Open_Excel_Workbook_Blank():
    Excel_Workbook = xlrd.open_workbook()
    return Excel_Workbook
# ---
def XLRD_Open_Excel_Workbook_Default(Fldr_Name,Option):
    Excel_Workbook = xlrd.open_workbook(Get_File_Path(Fldr_Name,Option))
    return Excel_Workbook
# ---
def XLRD_Open_Excel_Workbook_Mannual(File_Path):
    Excel_Workbook = xlrd.open_workbook(File_Path)
    return Excel_Workbook
# ---
def OpenPyXL_Open_Excel_Workbook_Default(Fldr_Name,Option):
    Excel_Workbook = openpyxl.load_workbook(Get_File_Path(Fldr_Name,Option))
    return Excel_Workbook
# ---
def OpenPyXL_Open_Excel_Workbook_Mannual(File_Path):
    Excel_Workbook = openpyxl.load_workbook(File_Path)
    return Excel_Workbook
# ---
# {Read-Open}-Excel_Workbook-Sheet
def XLRD_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Index):
    Excel_Workbook_Worksheet = Excel_Workbook.sheet_by_index(Index)    
    return Excel_Workbook_Worksheet
# ---
def OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,Index):
    Excel_Workbook_Worksheet = Excel_Workbook.worksheets[Index]    
    return Excel_Workbook_Worksheet

def OpenPyXL_Open_Excel_Workbook_Worksheet_ByName(Excel_Workbook,Name):
    Excel_Workbook_Worksheets = Excel_Workbook.sheetnames
    Excel_Workbook_Worksheet = Excel_Workbook[Excel_Workbook_Worksheets[int(Excel_Workbook_Worksheets.index(str(Name)))]]
    return Excel_Workbook_Worksheet

def OpenPyXL_Open_Excel_Workbook_Worksheet_ByActive(Excel_Workbook):
    Excel_Workbook_Worksheet = Excel_Workbook.active    
    return Excel_Workbook_Worksheet
# ---
# {Read}-Excel_Workbook-Len
def OpenPyXL_Read_Excel_Workbook_Length(Excel_Workbook):
    return len(Excel_Workbook.sheetnames)
# {Read}-Excel_Workbook-Sheet
# print(XLRD_Read_Excel_Workbook_Worksheet_Content(XLRD_Open_Excel_Workbook_Worksheet_ByIndex(XLRD_Open_Excel_Workbook_Default("In","100"),0),[0,0],[0,0]))
def XLRD_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,Index_Start,Index_End):
    if Index_Start[0]==Index_End[0] and Index_Start[1]==Index_End[1]:
        Point = []
        Point.append(str(Excel_Workbook_Worksheet.cell(Index_Start[0],Index_Start[1]).value))
        return Point
    if Index_Start[0]==Index_End[0] and Index_Start[1]!=Index_End[1]:
        Row_List = []
        for x in range(Index_Start[1],Index_End[1]+1):
            Row_List.append(str(Excel_Workbook_Worksheet.cell(Index_Start[0],x).value))
        return Row_List
    if Index_Start[1]==Index_End[1] and Index_Start[0]!=Index_End[0]:
        Col_List = []
        for y in range(Index_Start[0],Index_End[0]+1):
            Col_List.append(str(Excel_Workbook_Worksheet.cell(y,Index_Start[1]).value))
        return Col_List
    if Index_Start[1]!=Index_End[1] and Index_Start[0]!=Index_End[0]:
        Matrix = []
        for y in range(Index_Start[0],Index_End[0]+1):
            Matrix_Row_Temp=[]
            for x in range(Index_Start[1],Index_End[1]+1):
                Matrix_Row_Temp.append(str(Excel_Workbook_Worksheet.cell(y,x).value))
            Matrix.append(Matrix_Row_Temp)
        return Matrix
# ---
def OpenPyXL_Read_Excel_Workbook_Worksheet_Content(Excel_Workbook_Worksheet,Index_Start,Index_End):
    if Index_Start[0]==Index_End[0] and Index_Start[1]==Index_End[1]:
        Point = []
        Point.append(str(Excel_Workbook_Worksheet.cell(row=Index_Start[0],column=Index_Start[1]).value))
        return Point
    if Index_Start[0]==Index_End[0] and Index_Start[1]!=Index_End[1]:
        Row_List = []
        for x in range(Index_Start[1],Index_End[1]+1):
            Row_List.append(str(Excel_Workbook_Worksheet.cell(row=Index_Start[0],column=x).value))
        return Row_List
    if Index_Start[1]==Index_End[1] and Index_Start[0]!=Index_End[0]:
        Col_List = []
        for y in range(Index_Start[0],Index_End[0]+1):
            Col_List.append(str(Excel_Workbook_Worksheet.cell(row=y,column=Index_Start[1]).value))
        return Col_List
    if Index_Start[1]!=Index_End[1] and Index_Start[0]!=Index_End[0]:
        Matrix = []
        for y in range(Index_Start[0],Index_End[0]+1):
            Matrix_Row_Temp=[]
            for x in range(Index_Start[1],Index_End[1]+1):
                Matrix_Row_Temp.append(str(Excel_Workbook_Worksheet.cell(row=y,column=x).value))
            Matrix.append(Matrix_Row_Temp)
        return Matrix
# ---
def OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Workbook_Worksheet):
    Row_Total_Occupied = 0
    Col_Total_Occupied = 0
    Is_Row_Empty = False
    Is_Col_Empty = False
    for y in range(1,Excel_Workbook_Worksheet.max_row+2):
        Temp_Col_Count_Empty = 0
        for x in range(1,Excel_Workbook_Worksheet.max_column+2):
            if Excel_Workbook_Worksheet.cell(row = y, column=x).value == None:
                Temp_Col_Count_Empty = Temp_Col_Count_Empty + 1 
        if Temp_Col_Count_Empty == Excel_Workbook_Worksheet.max_column :
            Row_Total_Occupied = y-1
            break
    for x in range(1,Excel_Workbook_Worksheet.max_column+2):
        Temp_Row_Count_Empty = 0
        for y in range(1,Excel_Workbook_Worksheet.max_row+2):
            if Excel_Workbook_Worksheet.cell(row = y, column=x).value == None:
                Temp_Row_Count_Empty = Temp_Row_Count_Empty + 1 
        if Temp_Row_Count_Empty == Excel_Workbook_Worksheet.max_row :
            Col_Total_Occupied = x-1
            break
    return [Row_Total_Occupied,Col_Total_Occupied]
              
    # return [Excel_Workbook_Worksheet.max_row,Excel_Workbook_Worksheet.max_column]
# ---
# ---
# Example-XLRD-Read
# print(XLRD_Read_Excel_Workbook_Worksheet_Content(XLRD_Open_Excel_Workbook_Worksheet_ByIndex(XLRD_Open_Excel_Workbook_Default("In","100"),0),[0,0],[0,0]))

# Example-OpenPyXL-Read
# print(OpenPyXL_Read_Excel_Workbook_Worksheet_Content(OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(OpenPyXL_Open_Excel_Workbook_Default("In","100"),0),[1,1],[1,1]))

# Example-XLWT-Write
# Excel_Workbook,Excel_Workbook_Worksheet = XLWT_Create_Excel_Workbook_Worksheet_Default(XLWT_Create_Excel_Workbook_Blank())
# Excel_Workbook_Worksheet = XLWT_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[0,0],['A','B'],"H")
# Excel_Workbook = XLWT_Create_Excel_FromWorkbook(Excel_Workbook,Get_File_Path("Out","100"))

# Example-XLSXWriter-Write
# Excel_Workbook = XLSXWriter_Create_Excel_Workbook_FromPath_Mannual(Get_File_Path("Out","100"))
# Excel_Workbook_Worksheet = XLSXWriter_Create_Excel_Workbook_Worksheet_Mannual(Excel_Workbook,"Sheet_1")
# Excel_Workbook_Worksheet = XLSXWriter_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[0,0],[1,2,3,4],None,"H")
# Excel_Workbook.close()

# Example-OpenPyXL-Write
# Excel_Workbook = XLWT_Create_Excel_Workbook_Blank()
# Excel_Workbook_Worksheet = OpenPyXL_Write_Excel_Workbook_Worksheet(OpenPyXL_Create_Excel_Workbook_Worksheet_Mannual(Excel_Workbook,0,"Sheet1"),[1,1],[1,6,7,8],None,"V")
# OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,Get_File_Path("Out","100"))

# Example-XLSXWriter-Update
# Excel_Workbook_Worksheet = XLSXWriter_Write_Excel_Workbook_Worksheet(Excel_Workbook_Worksheet,[0,0],[5,2,3,4],None,"H")

# Example-OpenPyXL-Update
# Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(Get_File_Path("Out","100"))
# Excel_Workbook_Worksheet = OpenPyXL_Write_Excel_Workbook_Worksheet(OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,0),[1,1],[5],None,"V")
# OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,Get_File_Path("Out","100"))

# Example-Other
# print (wb.sheetnames)
# ws = wb.worksheets[0]


# ws = OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(OpenPyXL_Open_Excel_Workbook_Mannual("S:/Sahu_Group/A1_Plan/B1_EndPt/D1_Identity.xlsx"),2)
# print(ws.max_row)
# print(ws.max_column)


# print(OpenPyXL_Read_Excel_Workbook_Worksheet_Content(OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(OpenPyXL_Open_Excel_Workbook_Default("In","100"),0),[1,1],[1,1]))
# Open

# Example-OpenPyXL-Read-Workbook-Worksheet-Length
# print(OpenPyXL_Read_Excel_Workbook_Length(OpenPyXL_Open_Excel_Workbook_Mannual("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_MngtTimeWork/A1_In/A1_In_1.xlsx")))
# Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_MngtTimeWork/A1_In/A1_In_1.xlsx")
# print(OpenPyXL_Read_Excel_Workbook_Length(Excel_Workbook))
# Excel_Worksheet = OpenPyXL_Open_Excel_Workbook_Worksheet_ByActive(Excel_Workbook)
# print(OpenPyXL_Read_Excel_Workbook_Worksheet_Length(Excel_Worksheet)[1])

# OpenPyXL_Update_Excel_SerialDeSerialCellMatrix("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/LogicSerializer.xlsx",3,"Serialize")
# OpenPyXL_Update_Excel_KunchiKunchiCell("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/LogicSerializer.xlsx",2,"{}")
# OpenPyXL_Update_Excel_SerialDeSerialCellMatrix("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/LogicSerializer.xlsx",4,"Serialize")

# OpenPyXL_Update_Excel_FillDeFillCellMatrix("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/TableFiller.xlsx",3,"Fill")

# OpenPyXL_Update_Excel_IDReplaceCellMatrix("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/IDReplacer.xlsx",[[0,1]])


# File_Path = "S:/Sahu_Group/B1_Xport/A1_Excel/C1_Fuschia/D2L1_MyWork.xlsx"
# Excel_Workbook = OpenPyXL_Open_Excel_Workbook_Mannual(File_Path)
# Excel_Workbook_Worksheet = OpenPyXL_Write_Excel_Workbook_Worksheet(OpenPyXL_Open_Excel_Workbook_Worksheet_ByIndex(Excel_Workbook,0),[1,1],[5],None,"V")
# OpenPyXL_Save_Excel_Workbook_FromPath_Mannual(Excel_Workbook,File_Path)

# OpenPyXL_Update_Excel_SerialDeSerialCellMatrix("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/LogicSerializer.xlsx",5,"Serialize")
# OpenPyXL_TableDescionSerialDeSerial("S:/Sahu_Group/C3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/TableDescionSerialDeSerial.xlsx",0,"Serialize")

OpenPyXL_Update_Excel_SerialDeSerialCellMatrix("S:/Sahu_Group/B3_Data/D2_Raw_FldrFile/C4_RschngCopn/D1_MyProduction_IT/MyDevSyb/D2_Prjct_Atmn_Excel/A1_In/RelationSerializer.xlsx",0)

# print(Get_SerialToDeserial("{}-{}-{}"))
# print(Get_SerialToDeserial("Cy-FJ-{1999-2099}"))